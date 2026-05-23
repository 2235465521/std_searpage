"""标准数据分析：按行政区划、类别、年度统计与对比。"""
import io
import re
import time
from collections import defaultdict

from datetime import date

from django.db import connection
from django.db.models import Count, Max, Q
from django.db.models.functions import ExtractYear

from .redis_circuit import safe_cache_get, safe_cache_set
from .models import AreaDict, StdUnitRelation, UnitDict, ViewStdFull
from .std_type_util import normalize_std_type_code

_SUMMARY_CACHE = {}
_SUMMARY_CACHE_TTL = 600
_INDEX_CACHE = {}
_INDEX_CACHE_TTL = 900

STD_TYPE_ORDER = ('GB', 'HB', 'DB', 'TB', 'ISO', 'IEC', 'IEEE')
STD_TYPE_LABELS = {
    'GB': '国家标准',
    'HB': '行业标准',
    'DB': '地方标准',
    'TB': '团体标准',
    'ISO': '国际标准',
    'IEC': '国际电工标准',
    'IEEE': '电气电子标准',
}
_SCOPE_ALIAS_TO_CODE = {
    'GB': '00',
    'HB': '01',
    'DB': '02',
    'TB': '03',
    '00': '00',
    '01': '01',
    '02': '02',
    '03': '03',
}


def normalize_std_scope(std_scope):
    if not std_scope:
        return []
    if isinstance(std_scope, (list, tuple, set)):
        raw = std_scope
    else:
        raw = str(std_scope).split(',')
    valid = {'00', '01', '02', '03'}
    result = []
    seen = set()
    for token in raw:
        key = (token or '').strip().upper()
        code = _SCOPE_ALIAS_TO_CODE.get(key)
        if code in valid and code not in seen:
            result.append(code)
            seen.add(code)
    return result


def latest_release_year():
    """库中有发布日期的标准里，取最新年份；无数据时退回当前公历年。"""
    max_date = (
        ViewStdFull.objects.filter(release_date__isnull=False)
        .aggregate(max_date=Max('release_date'))
        .get('max_date')
    )
    return max_date.year if max_date else date.today().year


def list_provinces():
    rows = (
        AreaDict.objects.exclude(province_name__isnull=True)
        .exclude(province_name='')
        .values_list('province_name', flat=True)
        .distinct()
        .order_by('province_name')
    )
    return sorted(set(rows))


def list_cities(province):
    if not province:
        return []
    rows = (
        AreaDict.objects.filter(province_name=province)
        .exclude(city_name__isnull=True)
        .exclude(city_name='')
        .values_list('city_name', flat=True)
        .distinct()
        .order_by('city_name')
    )
    return sorted(set(rows))


def list_counties(province, city):
    if not province or not city:
        return []
    rows = (
        AreaDict.objects.filter(province_name=province, city_name=city)
        .exclude(county_name__isnull=True)
        .exclude(county_name='')
        .values_list('county_name', flat=True)
        .distinct()
        .order_by('county_name')
    )
    return sorted(set(rows))


def _area_codes_for_region(province=None, city=None, county=None):
    qs = AreaDict.objects.all()
    if province:
        qs = qs.filter(province_name=province)
    if city:
        qs = qs.filter(city_name=city)
    if county:
        qs = qs.filter(county_name=county)
    return [c for c in qs.values_list('area_code', flat=True) if c]


def _db_prefixes_from_area_codes(area_codes):
    """地方标准号 DB + 行政区划数字前缀。"""
    prefixes = set()
    for code in area_codes:
        digits = re.sub(r'\D', '', str(code))
        if len(digits) >= 2:
            prefixes.add(digits[:2])
        if len(digits) >= 4:
            prefixes.add(digits[:4])
        if len(digits) >= 6:
            prefixes.add(digits[:6])
    return sorted(prefixes, key=len, reverse=True)


def _breakdown_level(province=None, city=None, county=None):
    """未选中的最细层级决定按下一级区划拆分统计。"""
    if county:
        return None
    if city:
        return 'county'
    if province:
        return 'city'
    return 'province'


def _list_breakdown_names(level, province=None, city=None):
    if level == 'province':
        return list_provinces()
    if level == 'city':
        return list_cities(province)
    if level == 'county':
        return list_counties(province, city)
    return []


def _kwargs_for_breakdown_item(level, name, province=None, city=None):
    if level == 'province':
        return {'province': name}
    if level == 'city':
        return {'province': province, 'city': name}
    if level == 'county':
        return {'province': province, 'city': city, 'county': name}
    return {}


def _year_label(year):
    return f'{int(year)}年' if year else ''


def _region_label(province=None, city=None, county=None, breakdown_level=None, year=None):
    year_tag = f'（{_year_label(year)}）' if year else ''
    if breakdown_level == 'province':
        return f'全国{year_tag}（按省）'
    if breakdown_level == 'city':
        return f'{province}{year_tag}（按市）'
    if breakdown_level == 'county':
        return f'{province} {city}{year_tag}（按县/区）'
    if county:
        return f'{province} {city} {county}{year_tag}'
    if city:
        return f'{province} {city}{year_tag}'
    if province:
        return f'{province}{year_tag}'
    return f'全国{year_tag}' if year else '全国'


def _apply_year_filter(queryset, year=None):
    if year is None:
        return queryset
    return queryset.filter(release_date__year=int(year))


def _apply_std_scope_filter(queryset, std_scope=None):
    std_scope = normalize_std_scope(std_scope)
    if not std_scope:
        return queryset
    return queryset.filter(std_type_no__in=std_scope)


def build_region_queryset(province=None, city=None, county=None, year=None, std_scope=None):
    """按省/市/县筛选标准（地方标准号 + 起草单位区划关联）。"""
    if not province and not city and not county:
        qs = _apply_year_filter(ViewStdFull.objects.all(), year)
        return _apply_std_scope_filter(qs, std_scope)

    area_codes = _area_codes_for_region(province, city, county)
    if not area_codes:
        return ViewStdFull.objects.none()

    region_q = Q()
    for prefix in _db_prefixes_from_area_codes(area_codes):
        region_q |= Q(std_type='DB', std_id__istartswith=f'DB{prefix}')

    unit_ids = UnitDict.objects.filter(area_code__in=area_codes).values_list('unit_id', flat=True)
    if unit_ids:
        base_ids = (
            StdUnitRelation.objects.filter(unit_id__in=unit_ids)
            .values_list('base_id', flat=True)
            .distinct()
        )
        region_q |= Q(id__in=base_ids)

    if not region_q:
        return ViewStdFull.objects.none()
    qs = _apply_year_filter(ViewStdFull.objects.filter(region_q), year)
    return _apply_std_scope_filter(qs, std_scope)


def _counts_by_type(queryset):
    rows = queryset.values('std_type', 'std_type_no').annotate(count=Count('id'))
    merged = defaultdict(int)
    for row in rows:
        code = normalize_std_type_code(row.get('std_type'), row.get('std_type_no')) or '其他'
        merged[code] += row['count']
    return _counts_dict_to_result(dict(merged))


def _counts_dict_to_result(counts):
    items = []
    total = 0
    for code in STD_TYPE_ORDER:
        n = counts.get(code, 0)
        if n:
            items.append({'std_type': code, 'label': STD_TYPE_LABELS.get(code, code), 'count': n})
            total += n
    for code, n in sorted(counts.items()):
        if code not in STD_TYPE_ORDER and n:
            items.append({'std_type': code, 'label': code, 'count': n})
            total += n
    return total, items


def _breakdown_region_field(level):
    return {'province': 'province_name', 'city': 'city_name', 'county': 'county_name'}[level]


def _build_breakdown_indexes(breakdown_level, province=None, city=None):
    """一次性加载区划与单位关联，供内存聚合（避免按区划重复扫全表）。"""
    index_key = f'{breakdown_level}:{province or ""}:{city or ""}'
    cached = _INDEX_CACHE.get(index_key)
    if cached and time.time() < cached['expires']:
        return cached['data']

    field = _breakdown_region_field(breakdown_level)
    area_qs = AreaDict.objects.exclude(**{f'{field}__isnull': True}).exclude(**{field: ''})
    if province:
        area_qs = area_qs.filter(province_name=province)
    if city:
        area_qs = area_qs.filter(city_name=city)

    region_to_codes = defaultdict(list)
    area_to_region = {}
    for row in area_qs.values('area_code', 'province_name', 'city_name', 'county_name'):
        code = row.get('area_code')
        region = row.get(field)
        if not code or not region:
            continue
        region_to_codes[region].append(code)
        area_to_region[code] = region

    region_to_prefixes = {
        region: _db_prefixes_from_area_codes(codes)
        for region, codes in region_to_codes.items()
    }

    unit_to_region = {}
    if area_to_region:
        code_list = list(area_to_region.keys())
        for uid, acode in UnitDict.objects.filter(area_code__in=code_list).values_list(
            'unit_id', 'area_code',
        ):
            region = area_to_region.get(acode)
            if region is not None:
                unit_to_region[uid] = region

    base_to_regions = defaultdict(set)
    if unit_to_region:
        uid_list = list(unit_to_region.keys())
        for base_id, unit_id in StdUnitRelation.objects.filter(unit_id__in=uid_list).values_list(
            'base_id', 'unit_id',
        ):
            region = unit_to_region.get(unit_id)
            if region:
                base_to_regions[base_id].add(region)

    data = (region_to_codes, region_to_prefixes, base_to_regions)
    _INDEX_CACHE[index_key] = {'data': data, 'expires': time.time() + _INDEX_CACHE_TTL}
    return data


def _build_prefix_to_regions(region_to_prefixes):
    """前缀 -> 区划集合，地方标准号 O(1) 查找。"""
    mapping = defaultdict(set)
    for region, prefixes in region_to_prefixes.items():
        for prefix in prefixes:
            mapping[prefix].add(region)
    return mapping


def _match_db_regions_fast(std_id, prefix_to_regions):
    if not std_id:
        return set()
    sid = str(std_id).upper()
    if not sid.startswith('DB'):
        return set()
    digits = re.sub(r'\D', '', sid[2:])
    matched = set()
    for length in (6, 4, 2):
        if len(digits) >= length:
            matched.update(prefix_to_regions.get(digits[:length], ()))
    return matched


def _chunked(iterable, size):
    buf = []
    for item in iterable:
        buf.append(item)
        if len(buf) >= size:
            yield buf
            buf = []
    if buf:
        yield buf


def _scope_totals(province=None, city=None, county=None, year=None, std_scope=None):
    """范围合计：全国用单次聚合，有区划筛选时才走关联查询。"""
    if not province and not city and not county:
        qs = _apply_year_filter(ViewStdFull.objects.all(), year)
        return _counts_by_type(_apply_std_scope_filter(qs, std_scope))
    return _counts_by_type(build_region_queryset(province, city, county, year, std_scope=std_scope))


def _cache_key(prefix, province, city, county, year=None, std_scope=None):
    scope_key = ','.join(normalize_std_scope(std_scope))
    return f'{prefix}:v5:{province or ""}:{city or ""}:{county or ""}:{year or ""}:{scope_key}'


def _cache_get(key):
    result = safe_cache_get(key)
    if result is not None:
        return result
    entry = _SUMMARY_CACHE.get(key)
    if entry and time.time() < entry['expires']:
        return entry['data']
    return None


def _cache_set(key, data):
    _SUMMARY_CACHE[key] = {'data': data, 'expires': time.time() + _SUMMARY_CACHE_TTL}
    safe_cache_set(key, data, _SUMMARY_CACHE_TTL)


def clear_summary_cache():
    _SUMMARY_CACHE.clear()
    _INDEX_CACHE.clear()


def _all_breakdown_region_names(breakdown_level, province=None, city=None):
    if breakdown_level == 'province':
        return list_provinces()
    if breakdown_level == 'city' and province:
        return list_cities(province)
    if breakdown_level == 'county' and province and city:
        return list_counties(province, city)
    return []


def _aggregate_linked_breakdown_sql(
    breakdown_level,
    province=None,
    city=None,
    year=None,
    std_scope=None,
):
    """单位关联类标准：数据库侧 DISTINCT + GROUP BY，避免 Python 逐行扫全表。"""
    region_field = _breakdown_region_field(breakdown_level)
    v_conds = ['(v.std_type IS NULL OR v.std_type <> %s)']
    params = ['DB']

    if year is not None:
        v_conds.append('v.release_date IS NOT NULL AND YEAR(v.release_date) = %s')
        params.append(int(year))

    scope_codes = normalize_std_scope(std_scope)
    non_db_scope = [c for c in scope_codes if c != '02']
    if scope_codes:
        if not non_db_scope:
            return {}
        placeholders = ','.join(['%s'] * len(non_db_scope))
        v_conds.append(f'v.std_type_no IN ({placeholders})')
        params.extend(non_db_scope)

    ad_conds = [
        f'ad.{region_field} IS NOT NULL',
        f"ad.{region_field} <> ''",
    ]
    if province:
        ad_conds.append('ad.province_name = %s')
        params.append(province)
    if city and breakdown_level in ('city', 'county'):
        ad_conds.append('ad.city_name = %s')
        params.append(city)

    where_sql = ' AND '.join(v_conds + ad_conds)
    sql = f"""
        SELECT region_name, std_type, std_type_no, COUNT(*) AS cnt
        FROM (
            SELECT DISTINCT v.id AS base_id, ad.{region_field} AS region_name,
                   v.std_type, v.std_type_no
            FROM view_std_full v
            INNER JOIN std_unit_relation sur ON sur.base_id = v.id
            INNER JOIN unit_dict ud ON ud.unit_id = sur.unit_id
            INNER JOIN area_dict ad ON ad.area_code = ud.area_code
            WHERE {where_sql}
        ) AS paired
        GROUP BY region_name, std_type, std_type_no
    """

    region_counts = defaultdict(lambda: defaultdict(int))
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        for region_name, std_type, std_type_no, cnt in cursor.fetchall():
            if not region_name:
                continue
            code = normalize_std_type_code(std_type, std_type_no) or '其他'
            region_counts[region_name][code] += int(cnt or 0)
    return region_counts


def _aggregate_db_breakdown(region_counts, prefix_to_regions, base_to_regions, year=None):
    """地方标准：按区划前缀过滤后再匹配，避免无差别扫描全部 DB 记录。"""
    if not prefix_to_regions:
        return region_counts

    prefixes = list(prefix_to_regions.keys())
    for prefix_chunk in _chunked(prefixes, 50):
        prefix_q = Q()
        for prefix in prefix_chunk:
            prefix_q |= Q(std_id__istartswith=f'DB{prefix}')
        db_qs = _apply_year_filter(
            ViewStdFull.objects.filter(std_type_no='02').filter(prefix_q),
            year,
        )
        for row in db_qs.values('id', 'std_id').iterator(chunk_size=5000):
            matched = set(base_to_regions.get(row['id'], ()))
            matched |= _match_db_regions_fast(row.get('std_id'), prefix_to_regions)
            if not matched:
                continue
            for region in matched:
                region_counts[region]['DB'] += 1
    return region_counts


def _totals_for_breakdown(province, city, county, year, std_scope):
    """分省明细请求复用已缓存的合计，避免重复聚合。"""
    fast_key = _cache_key('summary:totals', province, city, county, year, std_scope=std_scope)
    cached_fast = _cache_get(fast_key)
    if cached_fast is not None:
        return cached_fast['total'], cached_fast['by_type']
    return _scope_totals(province, city, county, year, std_scope=std_scope)


def _regional_summary_breakdown(
    breakdown_level,
    province=None,
    city=None,
    county=None,
    year=None,
    std_scope=None,
):
    std_scope = normalize_std_scope(std_scope)
    std_scope_set = set(std_scope)
    _, region_to_prefixes, base_to_regions = _build_breakdown_indexes(
        breakdown_level, province, city,
    )
    prefix_to_regions = _build_prefix_to_regions(region_to_prefixes)
    all_names = _all_breakdown_region_names(breakdown_level, province, city)
    if not all_names:
        all_names = sorted(region_to_prefixes.keys())
    else:
        all_names = sorted(set(all_names) | set(region_to_prefixes.keys()))
    region_counts = {name: defaultdict(int) for name in all_names}

    if not std_scope or (std_scope_set - {'02'}):
        linked_counts = _aggregate_linked_breakdown_sql(
            breakdown_level, province, city, year, std_scope=std_scope,
        )
        for region_name, type_counts in linked_counts.items():
            bucket = region_counts.setdefault(region_name, defaultdict(int))
            for code, count in type_counts.items():
                bucket[code] += count

    if not std_scope or '02' in std_scope_set:
        _aggregate_db_breakdown(region_counts, prefix_to_regions, base_to_regions, year=year)

    breakdown = []
    for name in all_names:
        counts = dict(region_counts.get(name, {}))
        total_item, _ = _counts_dict_to_result(counts)
        breakdown.append({
            'region_name': name,
            'total': total_item,
            'counts': counts,
        })

    breakdown.sort(key=lambda x: (-x['total'], x['region_name']))
    total, by_type = _totals_for_breakdown(province, city, county, year, std_scope)
    return breakdown, total, by_type


def regional_summary(
    province=None,
    city=None,
    county=None,
    year=None,
    std_scope=None,
    include_breakdown=True,
):
    std_scope = normalize_std_scope(std_scope)
    breakdown_level = _breakdown_level(province, city, county)
    want_breakdown = bool(breakdown_level and include_breakdown)
    cache_prefix = 'summary' if want_breakdown else 'summary:totals'
    cache_key = _cache_key(cache_prefix, province, city, county, year, std_scope=std_scope)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    region_label = _region_label(province, city, county, breakdown_level, year)
    region_info = {
        'province': province or '',
        'city': city or '',
        'county': county or '',
        'year': int(year) if year else None,
        'label': region_label,
    }

    if want_breakdown:
        breakdown, total, by_type = _regional_summary_breakdown(
            breakdown_level, province, city, county, year, std_scope=std_scope,
        )
        result = {
            'region': region_info,
            'breakdown_level': breakdown_level,
            'breakdown': breakdown,
            'breakdown_pending': False,
            'total': total,
            'by_type': by_type,
            'std_scope': std_scope,
        }
    else:
        total, by_type = _scope_totals(province, city, county, year, std_scope=std_scope)
        result = {
            'region': region_info,
            'breakdown_level': breakdown_level,
            'breakdown': None,
            'breakdown_pending': bool(breakdown_level),
            'total': total,
            'by_type': by_type,
            'std_scope': std_scope,
        }

    _cache_set(cache_key, result)
    return result


def year_compare(province=None, city=None, county=None, year_a=None, year_b=None, std_scope=None):
    if year_a is None or year_b is None:
        raise ValueError('year_a and year_b are required')

    ck = _cache_key('yrcmp', province, city, county, year=f'{year_a}_{year_b}', std_scope=std_scope)
    cached = _cache_get(ck)
    if cached is not None:
        return cached

    qs = build_region_queryset(province, city, county, std_scope=std_scope).filter(
        release_date__isnull=False,
        release_date__year__in=[int(year_a), int(year_b)],
    )

    bucket = defaultdict(lambda: {int(year_a): 0, int(year_b): 0})
    rows = (
        qs.annotate(year=ExtractYear('release_date'))
        .values('std_type', 'std_type_no', 'year')
        .annotate(count=Count('id'))
    )
    for row in rows:
        st = normalize_std_type_code(row.get('std_type'), row.get('std_type_no')) or '其他'
        bucket[st][row['year']] += row['count']

    items = []
    all_types = list(STD_TYPE_ORDER) + [t for t in bucket if t not in STD_TYPE_ORDER]
    seen = set()
    for code in all_types:
        if code in seen:
            continue
        seen.add(code)
        ca = bucket[code].get(int(year_a), 0)
        cb = bucket[code].get(int(year_b), 0)
        if not ca and not cb:
            continue
        items.append({
            'std_type': code,
            'label': STD_TYPE_LABELS.get(code, code),
            'year_a': int(year_a),
            'year_b': int(year_b),
            'count_a': ca,
            'count_b': cb,
            'delta': cb - ca,
        })

    result = {
        'region': {
            'province': province or '',
            'city': city or '',
            'county': county or '',
            'label': _region_label(province, city, county, _breakdown_level(province, city, county)),
        },
        'year_a': int(year_a),
        'year_b': int(year_b),
        'std_scope': normalize_std_scope(std_scope),
        'items': items,
        'total_a': sum(i['count_a'] for i in items),
        'total_b': sum(i['count_b'] for i in items),
        'total_delta': sum(i['delta'] for i in items),
    }
    _cache_set(ck, result)
    return result


def year_range_stats(province=None, city=None, county=None, year_from=None, year_to=None, std_scope=None):
    if year_from is None or year_to is None:
        raise ValueError('year_from and year_to are required')

    yf, yt = int(year_from), int(year_to)
    if yf > yt:
        yf, yt = yt, yf
    if yf < 1900 or yt < 1900:
        raise ValueError('invalid year range')

    years = list(range(yf, yt + 1))
    ck = _cache_key('yrrng', province, city, county, year=f'{yf}_{yt}', std_scope=std_scope)
    cached = _cache_get(ck)
    if cached is not None:
        return cached

    qs = build_region_queryset(province, city, county, std_scope=std_scope).filter(
        release_date__isnull=False,
        release_date__year__gte=yf,
        release_date__year__lte=yt,
    )

    bucket = defaultdict(lambda: defaultdict(int))
    year_totals = defaultdict(int)
    rows = (
        qs.annotate(year=ExtractYear('release_date'))
        .values('std_type', 'std_type_no', 'year')
        .annotate(count=Count('id'))
    )
    for row in rows:
        year = int(row['year'])
        if year not in years:
            continue
        st = normalize_std_type_code(row.get('std_type'), row.get('std_type_no')) or '其他'
        count = row['count']
        bucket[st][year] += count
        year_totals[year] += count

    year_count = len(years)
    items = []
    all_types = list(STD_TYPE_ORDER) + [t for t in bucket if t not in STD_TYPE_ORDER]
    seen = set()
    for code in all_types:
        if code in seen:
            continue
        seen.add(code)
        year_counts = {y: bucket[code].get(y, 0) for y in years}
        total = sum(year_counts.values())
        if not total:
            continue
        items.append({
            'std_type': code,
            'label': STD_TYPE_LABELS.get(code, code),
            'year_counts': year_counts,
            'total': total,
            'avg_per_year': round(total / year_count, 1) if year_count else 0,
        })

    total = sum(year_totals.values())
    by_year = [
        {
            'year': y,
            'total': year_totals.get(y, 0),
            'by_type': [
                {
                    'std_type': code,
                    'label': STD_TYPE_LABELS.get(code, code),
                    'count': bucket[code].get(y, 0),
                }
                for code in all_types
                if bucket[code].get(y, 0)
            ],
        }
        for y in years
    ]

    by_type = []
    for item in items:
        share_pct = round(item['total'] * 100 / total, 1) if total else 0
        by_type.append({
            'std_type': item['std_type'],
            'label': item['label'],
            'count': item['total'],
            'avg_per_year': item['avg_per_year'],
            'share_pct': share_pct,
        })

    result = {
        'region': {
            'province': province or '',
            'city': city or '',
            'county': county or '',
            'label': _region_label(province, city, county, _breakdown_level(province, city, county)),
        },
        'year_from': yf,
        'year_to': yt,
        'years': years,
        'year_count': year_count,
        'std_scope': normalize_std_scope(std_scope),
        'by_year': by_year,
        'by_type': by_type,
        'items': items,
        'total': total,
        'avg_per_year': round(total / year_count, 1) if year_count else 0,
    }
    _cache_set(ck, result)
    return result


def build_excel_workbook(summary, compare=None, year_range=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws1 = wb.active
    ws1.title = '区域统计'

    region = summary['region']['label']
    ws1.append(['分析区域', region])
    ws1.append(['标准总数', summary['total']])
    ws1.append([])

    breakdown = summary.get('breakdown') or []
    if breakdown:
        level = summary.get('breakdown_level') or ''
        level_label = {'province': '省', 'city': '市', 'county': '县/区'}.get(level, '区划')
        header = [level_label, '合计'] + [
            STD_TYPE_LABELS.get(c, c) for c in STD_TYPE_ORDER
        ]
        ws1.append(header)
        for cell in ws1[ws1.max_row]:
            cell.font = Font(bold=True)
        for item in breakdown:
            counts = item.get('counts') or {}
            ws1.append(
                [item['region_name'], item['total']]
                + [counts.get(c, 0) for c in STD_TYPE_ORDER]
            )
        ws1.append([])
        ws1.append(['合计', summary['total']] + [
            sum((i.get('counts') or {}).get(c, 0) for i in breakdown)
            for c in STD_TYPE_ORDER
        ])
    else:
        ws1.append(['类别', '类型代码', '数量'])
        for row in summary['by_type']:
            ws1.append([row['label'], row['std_type'], row['count']])
        for cell in ws1[4]:
            cell.font = Font(bold=True)

    if compare:
        ws2 = wb.create_sheet('年度对比')
        ws2.append([
            '类别',
            '类型代码',
            f"{compare['year_a']}年发布数",
            f"{compare['year_b']}年发布数",
            '增减',
        ])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for row in compare['items']:
            ws2.append([
                row['label'],
                row['std_type'],
                row['count_a'],
                row['count_b'],
                row['delta'],
            ])
        ws2.append([])
        ws2.append([
            '合计',
            '',
            compare['total_a'],
            compare['total_b'],
            compare['total_delta'],
        ])

    if year_range:
        ws3 = wb.create_sheet('年段统计')
        yf = year_range['year_from']
        yt = year_range['year_to']
        years = year_range.get('years') or []
        ws3.append(['分析区域', year_range['region']['label']])
        ws3.append(['年段', f'{yf}—{yt}'])
        ws3.append(['段内合计', year_range['total']])
        ws3.append(['年均发布', year_range['avg_per_year']])
        ws3.append([])

        trend_header = ['年份', '发布合计'] + [
            STD_TYPE_LABELS.get(c, c) for c in STD_TYPE_ORDER
        ]
        ws3.append(trend_header)
        for cell in ws3[ws3.max_row]:
            cell.font = Font(bold=True)
        for row in year_range.get('by_year') or []:
            counts = {item['std_type']: item['count'] for item in (row.get('by_type') or [])}
            ws3.append(
                [row['year'], row['total']]
                + [counts.get(c, 0) for c in STD_TYPE_ORDER]
            )
        ws3.append([])
        ws3.append(['类别汇总'])
        ws3.append(['类别', '类型代码', '段内合计', '年均发布', '占比(%)'])
        for cell in ws3[ws3.max_row]:
            cell.font = Font(bold=True)
        for row in year_range.get('by_type') or []:
            ws3.append([
                row['label'],
                row['std_type'],
                row['count'],
                row['avg_per_year'],
                row['share_pct'],
            ])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
            for cell in col:
                cell.alignment = Alignment(vertical='center', wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
