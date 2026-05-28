"""起草单位检索：按标准内起草单位排位（支持单值与范围）筛选。"""
from collections import defaultdict
from io import BytesIO
import re

from django.core.cache import cache
from django.db import connection
from django.db.models import Exists, OuterRef, Q as models_Q
from .analytics import _area_codes_for_region, latest_release_year
from .models import AreaDict, StdExtendH, StdUnitRelation, UnitDict, ViewStdFull
from .services import extract_drafter_names
from .std_type_util import normalize_std_type_code

COUNT_TIER_RANK1 = 'eq1'
COUNT_TIER_RANK_2_4 = 'range_2_4'
COUNT_TIER_RANK_GTE4 = 'gte4'
VALID_COUNT_TIERS = {COUNT_TIER_RANK1, COUNT_TIER_RANK_2_4, COUNT_TIER_RANK_GTE4}

_CACHE_TTL = 600
_SCOPE_ALIAS_TO_CODE = {
    'GB': '00',
    'HB': '01',
    'DB': '02',
    'TB': '03',
    '00': '00',
    '01': '01',
    '02': '02',
    '03': '03',
}
VALID_STD_SCOPE_CODES = {'00', '01', '02', '03'}


def normalize_std_scope(std_scope):
    """归一化标准类别筛选，返回有序编码 list，如 ['00','02']。"""
    if not std_scope:
        return []
    if isinstance(std_scope, (list, tuple, set)):
        raw = std_scope
    else:
        raw = str(std_scope).split(',')
    normalized = []
    seen = set()
    for token in raw:
        key = (token or '').strip().upper()
        code = _SCOPE_ALIAS_TO_CODE.get(key)
        if code in VALID_STD_SCOPE_CODES and code not in seen:
            normalized.append(code)
            seen.add(code)
    return normalized


def normalize_count_tier(tier):
    tier = (tier or '').strip() or COUNT_TIER_RANK1
    return tier if tier in VALID_COUNT_TIERS else COUNT_TIER_RANK1


def _legacy_tier_to_range(tier):
    tier = normalize_count_tier(tier)
    if tier == COUNT_TIER_RANK1:
        return 1, 1
    if tier == COUNT_TIER_RANK_2_4:
        return 2, 3
    return 4, None


def _clean_rank_query(rank_query):
    if rank_query is None:
        return ''
    query = str(rank_query).strip()
    if not query:
        return ''
    # 允许“第7位/7名”等自然输入，只保留数字与比较/范围符号
    query = (
        query.replace('第', '')
        .replace('位', '')
        .replace('名', '')
        .replace(' ', '')
        .replace('，', ',')
        .replace('～', '-')
        .replace('—', '-')
        .replace('–', '-')
    )
    query = query.replace('到', '-').replace('至', '-')
    return query


def parse_rank_filter(rank_query=None, count_tier=None):
    """
    解析排位查询表达式，返回:
      {
        'min_rank': int,
        'max_rank': int|None,
        'query': str,   # 规范化后的查询表达式
        'label': str,   # 用于前端展示
      }
    支持示例: 7 / 2-4 / >=4 / >3 / <=10 / <5
    """
    cleaned = _clean_rank_query(rank_query)
    if not cleaned:
        min_rank, max_rank = _legacy_tier_to_range(count_tier)
        label = f'{min_rank}' if max_rank == min_rank else (
            f'{min_rank}-' if max_rank is None else f'{min_rank}-{max_rank}'
        )
        return {'min_rank': min_rank, 'max_rank': max_rank, 'query': label, 'label': label}

    if re.fullmatch(r'\d+', cleaned):
        rank = int(cleaned)
        if rank < 1:
            raise ValueError('排位必须是大于等于 1 的数字')
        return {'min_rank': rank, 'max_rank': rank, 'query': str(rank), 'label': str(rank)}

    m = re.fullmatch(r'(>=|<=|>|<)(\d+)', cleaned)
    if m:
        op, raw_num = m.groups()
        num = int(raw_num)
        if num < 1:
            raise ValueError('排位必须是大于等于 1 的数字')
        if op == '>=':
            return {'min_rank': num, 'max_rank': None, 'query': f'>={num}', 'label': f'>={num}'}
        if op == '>':
            return {'min_rank': num + 1, 'max_rank': None, 'query': f'>{num}', 'label': f'>{num}'}
        if op == '<=':
            return {'min_rank': 1, 'max_rank': num, 'query': f'<= {num}'.replace(' ', ''), 'label': f'<= {num}'.replace(' ', '')}
        if num <= 1:
            raise ValueError('排位范围为空，请调整查询条件')
        return {'min_rank': 1, 'max_rank': num - 1, 'query': f'<{num}', 'label': f'<{num}'}

    m = re.fullmatch(r'(\d+)-(\d+)', cleaned)
    if m:
        left, right = int(m.group(1)), int(m.group(2))
        if left < 1 or right < 1:
            raise ValueError('排位必须是大于等于 1 的数字')
        start, end = sorted((left, right))
        return {'min_rank': start, 'max_rank': end, 'query': f'{start}-{end}', 'label': f'{start}-{end}'}

    nums = [int(n) for n in re.findall(r'\d+', cleaned)]
    if len(nums) == 1 and nums[0] >= 1:
        rank = nums[0]
        return {'min_rank': rank, 'max_rank': rank, 'query': str(rank), 'label': str(rank)}
    if len(nums) >= 2:
        start, end = sorted((nums[0], nums[1]))
        if start < 1:
            raise ValueError('排位必须是大于等于 1 的数字')
        return {'min_rank': start, 'max_rank': end, 'query': f'{start}-{end}', 'label': f'{start}-{end}'}
    raise ValueError('排位格式无效，示例：7、2-4、>=4')


def _rank_in_filter(rank, rank_filter):
    if rank < rank_filter['min_rank']:
        return False
    max_rank = rank_filter['max_rank']
    return max_rank is None or rank <= max_rank


def _cache_key(year, rank_query, std_scope, province, city, county):
    std_scope_key = ','.join(std_scope or [])
    year_key = year if year is not None else 'all'
    return (
        f'drafter_rank_search:v8:{year_key}:{rank_query}:{std_scope_key}:'
        f'{province or ""}:{city or ""}:{county or ""}'
    )


def _region_unit_ids(area_codes):
    return set(
        UnitDict.objects.filter(area_code__in=area_codes).values_list('unit_id', flat=True)
    )


def _unit_in_region(unit_name, unit_id, region_unit_ids, area_codes):
    """起草单位是否在所选省/市/县（按单位库 area_code）。"""
    if unit_id:
        return unit_id in region_unit_ids
    name = (unit_name or '').strip()
    if not name:
        return False
    return UnitDict.objects.filter(unit_name=name, area_code__in=area_codes).exists()


def _unit_region_from_relation(rel):
    """单位在 unit_dict 中的注册省/市/县。"""
    if not rel.unit_id or not rel.unit:
        return None
    area_code = (rel.unit.area_code or '').strip()
    if not area_code:
        return None
    area = AreaDict.objects.filter(area_code=area_code).values(
        'province_name', 'city_name', 'county_name',
    ).first()
    if not area:
        return None
    return {
        'province': area.get('province_name') or '',
        'city': area.get('city_name') or '',
        'county': area.get('county_name') or '',
    }


def _build_relation_map(base_ids):
    """base_id -> [(rank, unit_name, unit_id|None), ...] 按 rank_order 排序，缺省则按顺序 1,2,3…"""
    relation_map = defaultdict(list)
    rows = (
        StdUnitRelation.objects.filter(base_id__in=base_ids)
        .select_related('unit')
        .order_by('base_id', 'rank_order', 'id')
    )
    for rel in rows:
        name = (rel.unit.unit_name or '').strip() if rel.unit_id else ''
        if not name:
            continue
        relation_map[rel.base_id].append(rel)

    ranked = {}
    for base_id, rels in relation_map.items():
        pairs = []
        for idx, rel in enumerate(rels):
            rank = rel.rank_order
            if rank is None or rank < 1:
                rank = idx + 1
            else:
                rank = int(rank)
            pairs.append((rank, rel.unit.unit_name.strip(), rel.unit_id))
        ranked[base_id] = pairs
    return ranked


def _ranked_names_from_text(row, extend_map):
    sources = []
    if row.drafter:
        sources.append(row.drafter)
    draft_unit = extend_map.get(row.id)
    if draft_unit:
        sources.append(draft_unit)
    names = extract_drafter_names(*sources)
    return [(i + 1, name, None) for i, name in enumerate(names)]


def _pick_units_for_tier(ranked_triples, rank_filter):
    return [
        (rank, name, unit_id)
        for rank, name, unit_id in ranked_triples
        if _rank_in_filter(rank, rank_filter)
    ]


def _ranked_triples_from_relations(rels):
    pairs = []
    for idx, rel in enumerate(rels):
        rank = rel.rank_order
        if rank is None or rank < 1:
            rank = idx + 1
        else:
            rank = int(rank)
        name = (rel.unit.unit_name or '').strip() if rel.unit_id and rel.unit else ''
        if name:
            pairs.append((rank, name, rel.unit_id))
    return pairs


def _make_match_item(row, picked, year, rank_filter):
    return {
        'std_id': row.std_id or '',
        'std_chinesename': row.std_chinesename or '',
        'std_type': normalize_std_type_code(row.std_type, row.std_type_no),
        'release_date': row.release_date.isoformat() if row.release_date else None,
        'ex_state': row.ex_state,
        'ranks': [r for r, _, _ in picked],
        'unit_names': [n for _, n, _ in picked],
        'unit_ids': [uid for _, _, uid in picked],
        'unit_name': picked[0][1],
        'year': int(year) if year is not None else None,
        'rank_query': rank_filter['query'],
    }


def _collect_relation_matches(year, rank_filter, std_scope, region_unit_ids=None, area_codes=None):
    """从 std_unit_relation 出发关联标准，避免按年扫全表标准。"""
    rel_qs = (
        StdUnitRelation.objects.select_related('unit', 'base')
        .filter(base__release_date__isnull=False)
        .order_by('base_id', 'rank_order', 'id')
    )
    if year is not None:
        rel_qs = rel_qs.filter(base__release_date__year=int(year))
    std_scope = normalize_std_scope(std_scope)
    if std_scope:
        rel_qs = rel_qs.filter(base__std_type_no__in=std_scope)
    if region_unit_ids is not None:
        rel_qs = rel_qs.filter(
            base_id__in=StdUnitRelation.objects.filter(
                unit_id__in=region_unit_ids,
            ).values('base_id'),
        )

    items = []
    current_base_id = None
    current_row = None
    current_rels = []

    def flush_base():
        nonlocal current_base_id, current_row, current_rels
        if current_base_id is None or current_row is None:
            return
        ranked_pairs = _ranked_triples_from_relations(current_rels)
        picked = _pick_units_for_tier(ranked_pairs, rank_filter)
        if region_unit_ids is not None:
            picked = [
                (r, n, uid)
                for r, n, uid in picked
                if _unit_in_region(n, uid, region_unit_ids, area_codes)
            ]
        if picked:
            items.append(_make_match_item(current_row, picked, year, rank_filter))
        current_rels = []

    for rel in rel_qs.iterator(chunk_size=3000):
        if rel.base_id != current_base_id:
            flush_base()
            current_base_id = rel.base_id
            current_row = rel.base
            current_rels = []
        current_rels.append(rel)
    flush_base()
    return items


def _collect_orphan_text_matches(
    year,
    rank_filter,
    std_scope,
    region_unit_ids=None,
    area_codes=None,
):
    """无单位关联记录的标准：仅对该子集做文本解析兜底。"""
    orphan_qs = (
        ViewStdFull.objects.annotate(
            has_rel=Exists(StdUnitRelation.objects.filter(base_id=OuterRef('id'))),
        )
        .filter(has_rel=False, release_date__isnull=False)
        .only(
            'id', 'std_id', 'std_chinesename', 'std_type', 'std_type_no',
            'release_date', 'drafter', 'ex_state',
        )
    )
    if year is not None:
        orphan_qs = orphan_qs.filter(release_date__year=int(year))
    std_scope = normalize_std_scope(std_scope)
    if std_scope:
        orphan_qs = orphan_qs.filter(std_type_no__in=std_scope)

    items = []
    batch_ids = []
    batch_rows = []

    def process_batch():
        if not batch_rows:
            return
        extend_map = dict(
            StdExtendH.objects.filter(base_id__in=batch_ids)
            .exclude(draft_unit__isnull=True)
            .exclude(draft_unit='')
            .values_list('base_id', 'draft_unit')
        )
        for row in batch_rows:
            ranked_pairs = _ranked_names_from_text(row, extend_map)
            if not ranked_pairs:
                continue
            picked = _pick_units_for_tier(ranked_pairs, rank_filter)
            if region_unit_ids is not None:
                picked = [
                    (r, n, uid)
                    for r, n, uid in picked
                    if _unit_in_region(n, uid, region_unit_ids, area_codes)
                ]
            if picked:
                items.append(_make_match_item(row, picked, year, rank_filter))
        batch_ids.clear()
        batch_rows.clear()

    for row in orphan_qs.iterator(chunk_size=2000):
        batch_ids.append(row.id)
        batch_rows.append(row)
        if len(batch_ids) >= 2000:
            process_batch()
    process_batch()
    return items


def _collect_matches(year, rank_filter, std_scope=None, province=None, city=None, county=None):
    area_codes = None
    region_unit_ids = None
    if province or city or county:
        area_codes = _area_codes_for_region(province, city, county)
        if not area_codes:
            return []
        region_unit_ids = _region_unit_ids(area_codes)

    items = _collect_relation_matches(
        year, rank_filter, std_scope,
        region_unit_ids=region_unit_ids,
        area_codes=area_codes,
    )
    items.extend(
        _collect_orphan_text_matches(
            year,
            rank_filter,
            std_scope,
            region_unit_ids=region_unit_ids,
            area_codes=area_codes,
        ),
    )
    items.sort(key=lambda x: (x['release_date'] or '', x['std_id']), reverse=True)
    return items


def _collect_analysis(items, rank_filter):
    std_type_counts = defaultdict(int)
    rank_counts = defaultdict(int)
    unit_total = 0

    for item in items:
        std_type = normalize_std_type_code(item.get('std_type')) or '未知'
        std_type_counts[std_type] += 1
        for rank in item.get('ranks') or []:
            rank_counts[int(rank)] += 1
        for name in item.get('unit_names') or []:
            name = (name or '').strip()
            if name:
                unit_total += 1

    std_type_rows = [
        {'std_type': k, 'count': v}
        for k, v in sorted(std_type_counts.items(), key=lambda x: (-x[1], x[0]))
    ]
    rank_rows = [
        {'rank': r, 'count': c}
        for r, c in sorted(rank_counts.items(), key=lambda x: x[0])
    ]
    return {
        'total': len(items),
        'unit_total': unit_total,
        'rank_filter': rank_filter['query'],
        'std_type_counts': std_type_rows,
        'rank_counts': rank_rows,
    }


def _resolve_region_breakdown_level(province=None, city=None, county=None):
    if not province:
        return 'province', '省'
    if not city:
        return 'city', '市'
    return 'county', '区/县'


def _collect_region_breakdown(items, province=None, city=None, county=None):
    level, label = _resolve_region_breakdown_level(province, city, county)
    if not items:
        return {'level': level, 'label': label, 'rows': []}
    selected_area_codes = set(_area_codes_for_region(province, city, county)) if (province or city or county) else set()

    unit_ids = set()
    fallback_names = set()
    for item in items:
        names = item.get('unit_names') or []
        ids = item.get('unit_ids') or []
        for idx, name in enumerate(names):
            unit_id = ids[idx] if idx < len(ids) else None
            if unit_id:
                unit_ids.add(unit_id)
            else:
                clean_name = (name or '').strip()
                if clean_name:
                    fallback_names.add(clean_name)

    id_to_area = {}
    if unit_ids:
        unit_qs = UnitDict.objects.filter(unit_id__in=unit_ids)
        if selected_area_codes:
            unit_qs = unit_qs.filter(area_code__in=selected_area_codes)
        for unit_id, area_code in unit_qs.values_list('unit_id', 'area_code'):
            if unit_id and area_code and unit_id not in id_to_area:
                id_to_area[unit_id] = area_code

    name_to_area = {}
    if fallback_names:
        name_qs = UnitDict.objects.filter(unit_name__in=fallback_names)
        if selected_area_codes:
            name_qs = name_qs.filter(area_code__in=selected_area_codes)
        rows = (
            name_qs.exclude(area_code__isnull=True)
            .exclude(area_code='')
            .values_list('unit_name', 'area_code')
        )
        for unit_name, area_code in rows:
            key = (unit_name or '').strip()
            if key and key not in name_to_area:
                name_to_area[key] = area_code

    area_codes = set(id_to_area.values()) | set(name_to_area.values())
    area_rows = AreaDict.objects.filter(area_code__in=area_codes).values(
        'area_code', 'province_name', 'city_name', 'county_name',
    ) if area_codes else []
    area_map = {row['area_code']: row for row in area_rows}

    counter = defaultdict(int)
    unknown_bucket = '未匹配区划'
    for item in items:
        names = item.get('unit_names') or []
        ids = item.get('unit_ids') or []
        for idx, name in enumerate(names):
            unit_id = ids[idx] if idx < len(ids) else None
            area_code = id_to_area.get(unit_id)
            if not area_code:
                area_code = name_to_area.get((name or '').strip())
            area = area_map.get(area_code)
            if not area:
                bucket = unknown_bucket
            elif level == 'province':
                bucket = (area.get('province_name') or '').strip() or unknown_bucket
            elif level == 'city':
                bucket = (area.get('city_name') or '').strip() or unknown_bucket
            else:
                bucket = (area.get('county_name') or '').strip() or unknown_bucket
            if bucket:
                counter[bucket] += 1

    rows = [{'name': k, 'count': v} for k, v in sorted(counter.items(), key=lambda x: (-x[1], x[0]))]
    return {'level': level, 'label': label, 'rows': rows}


def _query_all_items(
    year=None,
    count_tier=COUNT_TIER_RANK1,
    rank_query=None,
    std_scope=None,
    province=None,
    city=None,
    county=None,
):
    if year is not None:
        year = int(year)
    rank_filter = parse_rank_filter(rank_query=rank_query, count_tier=count_tier)
    std_scope = normalize_std_scope(std_scope)

    cache_key = _cache_key(year, rank_filter['query'], std_scope, province, city, county)
    all_items = cache.get(cache_key)
    if all_items is None:
        all_items = _collect_matches(year, rank_filter, std_scope, province, city, county)
        cache.set(cache_key, all_items, _CACHE_TTL)
    return all_items, year, rank_filter, std_scope


def search_units(
    keyword=None,
    province=None,
    city=None,
    county=None,
    year=None,
    count_tier=COUNT_TIER_RANK1,
    rank_query=None,
    std_scope=None,
    page=1,
    size=20,
    include_analysis=True,
):
    all_items, year, rank_filter, std_scope = _query_all_items(
        year=year,
        count_tier=count_tier,
        rank_query=rank_query,
        std_scope=std_scope,
        province=province,
        city=city,
        county=county,
    )

    total = len(all_items)
    page = max(1, page)
    size = max(1, min(size, 100))
    offset = (page - 1) * size
    items = all_items[offset:offset + size]

    if include_analysis:
        analysis = _collect_analysis(all_items, rank_filter)
        analysis['region_breakdown'] = _collect_region_breakdown(
            all_items, province=province, city=city, county=county,
        )
    else:
        analysis = None
    return total, items, year, rank_filter, std_scope, analysis


def build_export_data(
    province=None,
    city=None,
    county=None,
    year=None,
    count_tier=COUNT_TIER_RANK1,
    rank_query=None,
    std_scope=None,
):
    all_items, year, rank_filter, std_scope = _query_all_items(
        year=year,
        count_tier=count_tier,
        rank_query=rank_query,
        std_scope=std_scope,
        province=province,
        city=city,
        county=county,
    )
    analysis = _collect_analysis(all_items, rank_filter)
    analysis['region_breakdown'] = _collect_region_breakdown(
        all_items, province=province, city=city, county=county,
    )
    return all_items, analysis, year, rank_filter, std_scope


def build_excel_workbook(items, analysis):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = '明细'
    ws.append(['标准号', '标准名称', '类型', '发布日期', '排位', '起草单位'])
    for item in items:
        ws.append([
            item.get('std_id') or '',
            item.get('std_chinesename') or '',
            item.get('std_type') or '',
            item.get('release_date') or '',
            '、'.join(str(x) for x in (item.get('ranks') or [])),
            '\n'.join(item.get('unit_names') or []),
        ])

    s2 = wb.create_sheet('统计')
    s2.append(['指标', '值'])
    s2.append(['命中标准数', analysis.get('total', 0)])
    s2.append(['命中起草单位数', analysis.get('unit_total', 0)])
    s2.append(['排位查询', analysis.get('rank_filter', '')])
    s2.append([])
    s2.append(['类型', '数量'])
    for row in analysis.get('std_type_counts') or []:
        s2.append([row.get('std_type') or '', row.get('count') or 0])
    s2.append([])
    s2.append(['排位', '命中次数'])
    for row in analysis.get('rank_counts') or []:
        s2.append([row.get('rank'), row.get('count') or 0])

    region_breakdown = analysis.get('region_breakdown') or {}
    s2.append([])
    s2.append([f"按{region_breakdown.get('label') or '区域'}分布", '命中次数'])
    for row in region_breakdown.get('rows') or []:
        s2.append([row.get('name') or '', row.get('count') or 0])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _date_sort_key(value):
    return value or '9999-12-31'


def find_first_lead_national_standard(province=None, city=None, county=None):
    """
    历史第一家牵头国家标准单位（std_type_no=00 且排位=1）。
    省/市/县筛选按单位区划匹配。
    用直接 DB 查询代替全量扫描，避免超时。
    """
    ck = f'first_lead:v2:{province or ""}:{city or ""}:{county or ""}'
    cached = cache.get(ck)
    if cached is not None:
        return cached if cached != '__NONE__' else None

    area_codes = None
    region_unit_ids = None
    if province or city or county:
        area_codes = _area_codes_for_region(province, city, county)
        if not area_codes:
            return None
        region_unit_ids = _region_unit_ids(area_codes)
        if not region_unit_ids:
            return None

    rel_qs = StdUnitRelation.objects.select_related('unit').filter(
        base__std_type_no='00',
        base__release_date__isnull=False,
    ).filter(
        models_Q(rank_order=1) | models_Q(rank_order__isnull=True),
    )
    if region_unit_ids is not None:
        rel_qs = rel_qs.filter(unit_id__in=region_unit_ids)

    rel_qs = rel_qs.order_by('base__release_date', 'base__std_id', 'id')

    for rel in rel_qs.iterator(chunk_size=500):
        unit_name = (rel.unit.unit_name or '').strip() if rel.unit_id else ''
        if not unit_name:
            continue
        rank = rel.rank_order
        if rank is not None and rank != 1:
            continue
        std = rel.base
        if not std or not std.release_date:
            continue
        result = {
            'unit_name': unit_name,
            'std_id': std.std_id or '',
            'std_chinesename': std.std_chinesename or '',
            'std_type': normalize_std_type_code(std.std_type, std.std_type_no),
            'release_date': std.release_date.isoformat(),
            'rank': 1,
            'query_region': {
                'province': province or '',
                'city': city or '',
                'county': county or '',
            },
            'unit_region': _unit_region_from_relation(rel) or {},
        }
        cache.set(ck, result, _CACHE_TTL)
        return result
    cache.set(ck, '__NONE__', _CACHE_TTL)
    return None


def _find_first_participation_from_relations(unit_name, std_scope_codes):
    base_ids_and_ranks = list(
        StdUnitRelation.objects.filter(unit__unit_name=unit_name)
        .values_list('base_id', 'rank_order')
    )
    if not base_ids_and_ranks:
        return None

    rank_by_base = {}
    for base_id, rank_order in base_ids_and_ranks:
        rank = rank_order if rank_order and rank_order > 0 else 1
        if base_id not in rank_by_base or rank < rank_by_base[base_id]:
            rank_by_base[base_id] = rank

    qs = ViewStdFull.objects.filter(id__in=rank_by_base.keys(), release_date__isnull=False)
    if std_scope_codes:
        qs = qs.filter(std_type_no__in=std_scope_codes)
    row = qs.order_by('release_date', 'std_id').only(
        'id', 'std_id', 'std_chinesename', 'std_type', 'std_type_no', 'release_date',
    ).first()
    if not row:
        return None
    return {
        'unit_name': unit_name,
        'rank': int(rank_by_base.get(row.id, 1)),
        'std_id': row.std_id or '',
        'std_chinesename': row.std_chinesename or '',
        'std_type': normalize_std_type_code(row.std_type, row.std_type_no),
        'release_date': row.release_date.isoformat() if row.release_date else None,
        'match_source': 'relation',
    }


def _find_first_participation_from_text(unit_name, std_scope_codes):
    text_q = ViewStdFull.objects.filter(
        models_Q(drafter__icontains=unit_name)
        | models_Q(id__in=StdExtendH.objects.filter(draft_unit__icontains=unit_name).values_list('base_id', flat=True))
    ).filter(release_date__isnull=False)
    if std_scope_codes:
        text_q = text_q.filter(std_type_no__in=std_scope_codes)

    rows = text_q.order_by('release_date', 'std_id').only(
        'id', 'std_id', 'std_chinesename', 'std_type', 'std_type_no', 'release_date', 'drafter',
    )[:200]
    if not rows:
        return None

    base_ids = [r.id for r in rows]
    extend_map = dict(
        StdExtendH.objects.filter(base_id__in=base_ids).values_list('base_id', 'draft_unit')
    )
    target = unit_name.strip()
    for row in rows:
        ranked = _ranked_names_from_text(row, extend_map)
        for rank, name, _ in ranked:
            if (name or '').strip() == target:
                return {
                    'unit_name': target,
                    'rank': int(rank),
                    'std_id': row.std_id or '',
                    'std_chinesename': row.std_chinesename or '',
                    'std_type': normalize_std_type_code(row.std_type, row.std_type_no),
                    'release_date': row.release_date.isoformat() if row.release_date else None,
                    'match_source': 'text',
                }
    return None


def find_unit_first_participation(unit_name, std_scope=None):
    """
    查询某单位首次出现在起草单位名单中的标准（任意排位）。
    """
    target = (unit_name or '').strip()
    if not target:
        raise ValueError('unit_name 不能为空')

    scope_key = ','.join(normalize_std_scope(std_scope) or [])
    ck = f'unit_first_part:v1:{target}:{scope_key}'
    cached = cache.get(ck)
    if cached is not None:
        return cached if cached != '__NONE__' else None

    std_scope_codes = normalize_std_scope(std_scope)
    rel_hit = _find_first_participation_from_relations(target, std_scope_codes)
    text_hit = _find_first_participation_from_text(target, std_scope_codes)

    result = None
    if rel_hit and text_hit:
        rel_key = (_date_sort_key(rel_hit.get('release_date')), rel_hit.get('std_id') or '')
        text_key = (_date_sort_key(text_hit.get('release_date')), text_hit.get('std_id') or '')
        result = rel_hit if rel_key <= text_key else text_hit
    else:
        result = rel_hit or text_hit
    cache.set(ck, result if result is not None else '__NONE__', _CACHE_TTL)
    return result


_FIRST_PART_CACHE_TTL = 60 * 15


def _fetch_first_participation_by_region_sql(area_codes, std_scope_codes):
    """
    单次 SQL：先按单位聚合最早参与年份，再 JOIN 回明细。
    避免 correlated subquery + 超大 unit_id IN 列表。
    """
    area_ph = ','.join(['%s'] * len(area_codes))
    scope_sub = ''
    scope_outer = ''
    params = list(area_codes)
    if std_scope_codes:
        scope_ph = ','.join(['%s'] * len(std_scope_codes))
        scope_sub = f' AND b2.std_type_no IN ({scope_ph})'
        scope_outer = f' AND b.std_type_no IN ({scope_ph})'
        params.extend(std_scope_codes)
    params.extend(area_codes)
    if std_scope_codes:
        params.extend(std_scope_codes)

    sql = f"""
        SELECT
            r.unit_id,
            u.unit_name,
            r.rank_order,
            fy.first_year,
            b.std_id,
            b.std_chinesename,
            b.std_type,
            b.std_type_no,
            b.release_date
        FROM std_unit_relation r
        INNER JOIN std_base b ON r.base_id = b.id
        INNER JOIN unit_dict u ON r.unit_id = u.unit_id
        INNER JOIN (
            SELECT r2.unit_id, MIN(YEAR(b2.release_date)) AS first_year
            FROM std_unit_relation r2
            INNER JOIN std_base b2 ON r2.base_id = b2.id
            INNER JOIN unit_dict u2 ON r2.unit_id = u2.unit_id
            WHERE u2.area_code IN ({area_ph})
              AND b2.release_date IS NOT NULL
              {scope_sub}
            GROUP BY r2.unit_id
        ) fy ON r.unit_id = fy.unit_id AND YEAR(b.release_date) = fy.first_year
        WHERE u.area_code IN ({area_ph})
          AND b.release_date IS NOT NULL
          {scope_outer}
        ORDER BY fy.first_year, u.unit_name, b.release_date, b.std_id, r.id
    """

    items = []
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        for (
            unit_id,
            unit_name,
            rank_order,
            first_year,
            std_id,
            std_chinesename,
            std_type,
            std_type_no,
            release_date,
        ) in cursor.fetchall():
            name = (unit_name or '').strip()
            if not name or not release_date:
                continue
            rank = int(rank_order) if rank_order and rank_order > 0 else 1
            items.append({
                'unit_id': unit_id,
                'unit_name': name,
                'rank': rank,
                'first_year': int(first_year or release_date.year),
                'std_id': std_id or '',
                'std_chinesename': std_chinesename or '',
                'std_type': normalize_std_type_code(std_type, std_type_no),
                'release_date': (
                    release_date.isoformat()
                    if hasattr(release_date, 'isoformat')
                    else str(release_date)
                ),
            })
    return items


def find_first_participation_by_region(province=None, city=None, county=None, std_scope=None):
    """
    查询某省/市/区内单位“首次参与标准编写”的记录列表。
    口径：
    - 按单位最早参与年份判定“首次参与年份”；
    - 返回该单位在首次参与年份内命中的全部标准记录（不去重）。
    """
    scope_key = ','.join(normalize_std_scope(std_scope) or [])
    ck = f'first_part:v2:{province or ""}:{city or ""}:{county or ""}:{scope_key}'
    cached = cache.get(ck)
    if cached is not None:
        return cached

    area_codes = _area_codes_for_region(province, city, county)
    if not area_codes:
        return []

    std_scope_codes = normalize_std_scope(std_scope)
    items = _fetch_first_participation_by_region_sql(area_codes, std_scope_codes)
    cache.set(ck, items, _FIRST_PART_CACHE_TTL)
    return items


FIRST_PART_EXPORT_MODES = {'detail', 'unit_summary', 'unit_first'}
FIRST_PART_EXPORT_SCOPES = {'all', 'page_range', 'page'}
_STD_SCOPE_LABELS = {'00': '国家标准', '01': '行业标准', '02': '地方标准', '03': '团体标准'}


def _parse_first_participation_year(value, field_name):
    if value is None or str(value).strip() == '':
        return None
    try:
        year = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'{field_name} 无效') from exc
    if year < 1900 or year > 2100:
        raise ValueError(f'{field_name} 无效')
    return year


def _apply_first_participation_export_filters(items, first_year_from=None, first_year_to=None, rank_filter=None):
    result = list(items or [])
    if first_year_from is not None:
        result = [row for row in result if int(row.get('first_year') or 0) >= first_year_from]
    if first_year_to is not None:
        result = [row for row in result if int(row.get('first_year') or 0) <= first_year_to]
    if rank_filter:
        result = [
            row for row in result
            if _rank_in_filter(int(row.get('rank') or 1), rank_filter)
        ]
    return result


def _transform_first_participation_export_items(items, export_mode='detail'):
    mode = (export_mode or 'detail').strip().lower()
    if mode not in FIRST_PART_EXPORT_MODES:
        raise ValueError('export_mode 无效，可选：detail、unit_summary、unit_first')
    if mode == 'detail':
        return list(items or [])

    grouped = defaultdict(list)
    for item in items or []:
        key = item.get('unit_id') or item.get('unit_name')
        grouped[key].append(item)

    if mode == 'unit_summary':
        summary = []
        for rows in grouped.values():
            rows_sorted = sorted(
                rows,
                key=lambda row: (_date_sort_key(row.get('release_date')), row.get('std_id') or ''),
            )
            first = rows_sorted[0]
            summary.append({
                'unit_name': first.get('unit_name') or '',
                'first_year': first.get('first_year'),
                'std_count': len(rows_sorted),
                'std_id': first.get('std_id') or '',
                'std_chinesename': first.get('std_chinesename') or '',
                'std_type': first.get('std_type') or '',
                'release_date': first.get('release_date') or '',
                'rank': first.get('rank'),
            })
        return sorted(
            summary,
            key=lambda row: (int(row.get('first_year') or 0), row.get('unit_name') or ''),
        )

    result = []
    for rows in grouped.values():
        rows_sorted = sorted(
            rows,
            key=lambda row: (_date_sort_key(row.get('release_date')), row.get('std_id') or ''),
        )
        result.append(rows_sorted[0])
    return sorted(
        result,
        key=lambda row: (
            int(row.get('first_year') or 0),
            row.get('unit_name') or '',
            _date_sort_key(row.get('release_date')),
        ),
    )


def _collect_first_participation_analysis(filtered_items, export_items, meta):
    unit_keys = set()
    first_year_counts = defaultdict(int)
    type_counts = defaultdict(int)
    rank_counts = defaultdict(int)
    for row in filtered_items or []:
        unit_keys.add(row.get('unit_id') or row.get('unit_name'))
        first_year = row.get('first_year')
        if first_year is not None:
            first_year_counts[int(first_year)] += 1
        std_type = row.get('std_type') or '未知'
        type_counts[std_type] += 1
        rank_counts[int(row.get('rank') or 1)] += 1

    return {
        'meta': meta,
        'filtered_record_total': len(filtered_items or []),
        'filtered_unit_total': len(unit_keys),
        'export_record_total': len(export_items or []),
        'first_year_counts': [
            {'first_year': year, 'count': count}
            for year, count in sorted(first_year_counts.items())
        ],
        'std_type_counts': [
            {'std_type': std_type, 'count': count}
            for std_type, count in sorted(type_counts.items(), key=lambda item: (-item[1], item[0]))
        ],
        'rank_counts': [
            {'rank': rank, 'count': count}
            for rank, count in sorted(rank_counts.items())
        ],
    }


def _prepare_first_participation_filtered(
    province=None,
    city=None,
    county=None,
    std_scope=None,
    first_year_from=None,
    first_year_to=None,
    rank_query=None,
):
    province = (province or '').strip()
    if not province:
        raise ValueError('请先选择省份')

    first_year_from = _parse_first_participation_year(first_year_from, 'first_year_from')
    first_year_to = _parse_first_participation_year(first_year_to, 'first_year_to')
    if first_year_from is not None and first_year_to is not None and first_year_from > first_year_to:
        raise ValueError('首次参与起始年不能大于结束年')

    std_scope = normalize_std_scope(std_scope)
    rank_filter = None
    rank_label = '不限'
    if rank_query is not None and str(rank_query).strip():
        rank_filter = parse_rank_filter(rank_query=rank_query)
        rank_label = rank_filter['label']

    items = find_first_participation_by_region(
        province=province,
        city=city,
        county=county,
        std_scope=std_scope,
    )
    filtered = _apply_first_participation_export_filters(
        items,
        first_year_from=first_year_from,
        first_year_to=first_year_to,
        rank_filter=rank_filter,
    )

    region_parts = [part for part in [province, city, county] if part]
    scope_labels = [_STD_SCOPE_LABELS.get(code, code) for code in (std_scope or [])]
    meta = {
        'province': province,
        'city': city or '',
        'county': county or '',
        'region_label': ' / '.join(region_parts),
        'std_scope': std_scope,
        'std_scope_label': '、'.join(scope_labels) if scope_labels else '全部类别',
        'first_year_from': first_year_from,
        'first_year_to': first_year_to,
        'first_year_label': _format_first_year_label(first_year_from, first_year_to),
        'rank_query': rank_label,
    }
    return filtered, meta


def query_first_participation_page(
    province=None,
    city=None,
    county=None,
    std_scope=None,
    first_year_from=None,
    first_year_to=None,
    rank_query=None,
    list_mode='detail',
    page=1,
    size=50,
):
    filtered, meta = _prepare_first_participation_filtered(
        province=province,
        city=city,
        county=county,
        std_scope=std_scope,
        first_year_from=first_year_from,
        first_year_to=first_year_to,
        rank_query=rank_query,
    )
    list_mode = (list_mode or 'detail').strip().lower()
    if list_mode not in FIRST_PART_EXPORT_MODES:
        raise ValueError('list_mode 无效，可选：detail、unit_summary、unit_first')

    transformed = _transform_first_participation_export_items(filtered, list_mode)
    try:
        page = max(1, int(page))
        size = min(200, max(1, int(size)))
    except (TypeError, ValueError):
        page, size = 1, 50
    total = len(transformed)
    start = (page - 1) * size
    return {
        'items': transformed[start:start + size],
        'total': total,
        'page': page,
        'size': size,
        'detail_total': len(filtered),
        'meta': {**meta, 'list_mode': list_mode},
    }


def _slice_first_participation_export_items(
    transformed,
    export_scope='all',
    page=1,
    page_from=None,
    page_to=None,
    size=50,
):
    try:
        size = min(200, max(1, int(size)))
    except (TypeError, ValueError):
        size = 50

    total = len(transformed or [])
    total_pages = max(1, (total + size - 1) // size) if total else 1
    scope = (export_scope or 'all').strip().lower()

    if scope == 'all':
        return transformed, 1, total_pages, size, total_pages

    if scope == 'page':
        start_page = end_page = max(1, int(page))
    else:
        try:
            start_page = max(1, int(page_from if page_from is not None else page))
            end_page = max(1, int(page_to if page_to is not None else start_page))
        except (TypeError, ValueError):
            raise ValueError('导出起止页码无效')

    if start_page > end_page:
        raise ValueError('导出起始页不能大于结束页')
    if start_page > total_pages:
        raise ValueError(f'导出起始页超出范围，共 {total_pages} 页')
    end_page = min(end_page, total_pages)

    start = (start_page - 1) * size
    end = end_page * size
    return (transformed or [])[start:end], start_page, end_page, size, total_pages


def build_first_participation_export_data(
    province=None,
    city=None,
    county=None,
    std_scope=None,
    first_year_from=None,
    first_year_to=None,
    rank_query=None,
    export_mode='detail',
    export_scope='all',
    page=1,
    page_from=None,
    page_to=None,
    size=50,
):
    filtered, meta = _prepare_first_participation_filtered(
        province=province,
        city=city,
        county=county,
        std_scope=std_scope,
        first_year_from=first_year_from,
        first_year_to=first_year_to,
        rank_query=rank_query,
    )

    export_mode = (export_mode or 'detail').strip().lower()
    export_scope = (export_scope or 'all').strip().lower()
    if export_scope not in FIRST_PART_EXPORT_SCOPES:
        raise ValueError('export_scope 无效，可选：all、page_range')

    transformed = _transform_first_participation_export_items(filtered, export_mode)
    export_items, start_page, end_page, page_size, total_pages = _slice_first_participation_export_items(
        transformed,
        export_scope=export_scope,
        page=page,
        page_from=page_from,
        page_to=page_to,
        size=size,
    )

    meta = {
        **meta,
        'export_mode': export_mode,
        'export_scope': export_scope,
        'page': start_page,
        'page_from': start_page,
        'page_to': end_page,
        'size': page_size,
        'total_pages': total_pages,
    }
    analysis = _collect_first_participation_analysis(filtered, export_items, meta)
    return export_items, analysis, meta


def _format_first_year_label(first_year_from, first_year_to):
    if first_year_from is not None and first_year_to is not None:
        return f'{first_year_from}—{first_year_to}'
    if first_year_from is not None:
        return f'>={first_year_from}'
    if first_year_to is not None:
        return f'<={first_year_to}'
    return '不限'


def _first_participation_export_mode_label(export_mode):
    return {
        'detail': '明细（首次参与年内全部标准）',
        'unit_summary': '单位汇总（每单位一行）',
        'unit_first': '单位首条（每单位最早一条）',
    }.get(export_mode, export_mode)


def _format_export_scope_label(meta):
    scope = (meta or {}).get('export_scope') or 'all'
    if scope == 'all':
        return '全部命中'
    page_from = meta.get('page_from') or meta.get('page') or 1
    page_to = meta.get('page_to') or page_from
    page_size = meta.get('size') or 50
    if page_from == page_to:
        return f'第 {page_from} 页（每页 {page_size} 条）'
    return f'第 {page_from}—{page_to} 页（每页 {page_size} 条）'


def build_first_participation_excel(export_items, analysis, export_mode='detail'):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = '明细'
    mode = (export_mode or 'detail').strip().lower()
    if mode == 'unit_summary':
        ws.append(['起草单位', '首次参与年份', '首次参与年内标准数', '最早标准编号', '最早标准名称', '类型', '最早发布日期', '最早参与排位'])
        for item in export_items:
            ws.append([
                item.get('unit_name') or '',
                item.get('first_year') or '',
                item.get('std_count') or 0,
                item.get('std_id') or '',
                item.get('std_chinesename') or '',
                item.get('std_type') or '',
                item.get('release_date') or '',
                item.get('rank') or '',
            ])
    else:
        ws.append(['起草单位', '首次参与年份', '标准编号', '标准名称', '类型', '发布日期', '参与排位'])
        for item in export_items:
            ws.append([
                item.get('unit_name') or '',
                item.get('first_year') or '',
                item.get('std_id') or '',
                item.get('std_chinesename') or '',
                item.get('std_type') or '',
                item.get('release_date') or '',
                item.get('rank') or '',
            ])

    meta = analysis.get('meta') or {}
    s2 = wb.create_sheet('统计')
    s2.append(['指标', '值'])
    s2.append(['统计范围', meta.get('region_label') or ''])
    s2.append(['标准类别', meta.get('std_scope_label') or ''])
    s2.append(['首次参与年份', meta.get('first_year_label') or ''])
    s2.append(['参与排位', meta.get('rank_query') or ''])
    s2.append(['导出粒度', _first_participation_export_mode_label(mode)])
    s2.append(['导出范围', _format_export_scope_label(meta)])
    s2.append(['筛选后记录数', analysis.get('filtered_record_total', 0)])
    s2.append(['筛选后单位数（去重）', analysis.get('filtered_unit_total', 0)])
    s2.append(['本次导出行数', analysis.get('export_record_total', 0)])
    s2.append([])
    s2.append(['首次参与年份', '记录数'])
    for row in analysis.get('first_year_counts') or []:
        s2.append([row.get('first_year'), row.get('count') or 0])
    s2.append([])
    s2.append(['标准类型', '记录数'])
    for row in analysis.get('std_type_counts') or []:
        s2.append([row.get('std_type') or '', row.get('count') or 0])
    s2.append([])
    s2.append(['参与排位', '记录数'])
    for row in analysis.get('rank_counts') or []:
        s2.append([row.get('rank'), row.get('count') or 0])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_unit_detail(std_id, year=None):
    year = int(year) if year else latest_release_year()
    row = ViewStdFull.objects.filter(std_id=std_id, release_date__year=year).first()
    if not row:
        return None

    relation_ranked = _build_relation_map([row.id])
    extend_map = dict(
        StdExtendH.objects.filter(base_id=row.id).values_list('base_id', 'draft_unit')
    )
    ranked_pairs = relation_ranked.get(row.id) or _ranked_names_from_text(row, extend_map)

    return {
        'std_id': row.std_id,
        'std_chinesename': row.std_chinesename,
        'ranked_units': [{'rank': r, 'unit_name': n} for r, n, _ in ranked_pairs],
        'unit_names': [n for _, n, _ in ranked_pairs],
        'year': year,
    }
